# -*- coding: utf-8 -*-
# 把 .xls (xlrd) 转成 .xlsx (openpyxl)：保留值、列宽、行高、合并单元格、表头加粗。
import sys, xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def cellval(cell):
    if cell.ctype == xlrd.XL_CELL_DATE:
        try: return xlrd.xldate.xldate_as_datetime(cell.value, 0).strftime('%Y-%m-%d')
        except Exception: return cell.value
    if cell.ctype in (xlrd.XL_CELL_NUMBER,):
        if cell.value == int(cell.value): return int(cell.value)
        return cell.value
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK): return None
    return cell.value

def col_letter(i):
    s = ''
    i += 1
    while i:
        i, r = divmod(i-1, 26)
        s = chr(65+r) + s
    return s

def is_empty(v):
    return v is None or (isinstance(v, str) and v.strip() == '')

src, dst = sys.argv[1], sys.argv[2]
rb = xlrd.open_workbook(src, formatting_info=True)
wb = Workbook()
wb.remove(wb.active)
for sh in rb.sheets():
    ws = wb.create_sheet(title=sh.name[:31])
    max_col_with_data = 0
    for r in range(min(sh.nrows, 300)):
        bold_row = (r == 0)
        for c in range(sh.ncols):
            v = cellval(sh.cell(r, c))
            if is_empty(v): continue
            cell = ws.cell(r+1, c+1, v)
            if bold_row:
                cell.font = Font(bold=True)
            if c + 1 > max_col_with_data:
                max_col_with_data = c + 1
    # 列宽：根据 formatting_info 里的列宽信息设置；无数据列也保留占位，避免转换后列数缩水
    for c in range(sh.ncols):
        ci = sh.colinfo_map.get(c)
        if ci and getattr(ci, 'width', 0) and ci.width > 0:
            cl = col_letter(c)
            ws.column_dimensions[cl].width = max(8, min(60, ci.width / 256.0 + 2))
            # 对没有任何数据的后备空列，在第 1 行写入空字符串，确保 ExcelJS 读回时能识别到该列
            if c + 1 > max_col_with_data:
                ws.cell(1, c+1, '')
    # 行高
    for r in range(min(sh.nrows, 300)):
        ri = sh.rowinfo_map.get(r)
        if ri and getattr(ri, 'height', 0) and ri.height > 0:
            ws.row_dimensions[r+1].height = max(10, min(200, ri.height / 20.0))
    # 合并单元格
    for (rlo, rhi, clo, chi) in sh.merged_cells:
        ws.merge_cells(start_row=rlo+1, start_column=clo+1, end_row=rhi, end_column=chi)
wb.save(dst)
print('converted', src, '->', dst)
